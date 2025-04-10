import requests
import re
from bs4 import BeautifulSoup
from urllib.parse import urljoin, urlparse
from openpyxl import Workbook, load_workbook
from rich.console import Console
from concurrent.futures import ThreadPoolExecutor, as_completed
import os
from rich.console import Console
from rich.panel import Panel
from rich.text import Text
from rich.prompt import Prompt
import tkinter as tk
from tkinter import filedialog

# API Keys (for Options 2 & 3)
SCRAPING_DOG_API_KEY = "Your ScrapingDog Api"
SERPAPI_KEY = "Your SerpApi"

console = Console()


if not os.path.exists("output"):
    os.makedirs("output")


INDUSTRY_KEYWORDS = [
    "software", "ai", "artificial intelligence", "cloud", "data", "e-learning", "cybersecurity",
    "healthcare", "fintech", "education", "devops", "iot", "blockchain", "logistics", "consulting"
]
NAME_PATTERN = re.compile(r"\b([A-Z][a-z]+(?:\s+[A-Z][a-z]+)+)\b")

# ----------------- Option 1: Full ICP Scraper -----------------
def extract_icp_from_website(domain):
    console.print(f"\n Scraping: [bold blue]{domain}[/bold blue]")
    try:
        url = f"https://{domain}"
        headers = {"User-Agent": "Mozilla/5.0"}
        res = requests.get(url, headers=headers, timeout=10)
        if res.status_code != 200:
            console.print(f"[red] Failed to access site: {domain} (Status {res.status_code})[/red]")
            return []
        soup = BeautifulSoup(res.text, 'html.parser')
        text = soup.get_text(separator=" ")

        # title
        title = soup.title.string.strip() if soup.title else "N/A"
        meta_desc = soup.find("meta", attrs={"name": "description"}) or soup.find("meta", attrs={"property": "og:description"})
        description = meta_desc["content"].strip() if meta_desc and meta_desc.get("content") else ""
        h1 = soup.find('h1')
        company_name = h1.get_text(strip=True) if h1 else title

        
        found_names = NAME_PATTERN.findall(text)
        contact_person = found_names[0] if found_names else "N/A"

        industry = "N/A"
        for keyword in INDUSTRY_KEYWORDS:
            if keyword.lower() in description.lower() or keyword.lower() in text.lower():
                industry = keyword.capitalize()
                break

        
        emails = list(set(re.findall(r"[a-zA-Z0-9_.+-]+@[a-zA-Z0-9-]+\.[a-zA-Z0-9-.]+", text)))

       
        phones_from_text = re.findall(r"\+?\d[\d\s\-().]{7,}", text)
        phones_from_tel = [a['href'].replace("tel:", "") for a in soup.find_all('a', href=True) if a['href'].startswith("tel:")]
        all_phones = set()
        for phone in phones_from_text + phones_from_tel:
            clean = re.sub(r"[^\d+]", "", phone)
            if 7 <= len(clean) <= 15:
                all_phones.add(clean)
        phones = list(all_phones) if all_phones else []

       
        whatsapp_links = [a['href'] for a in soup.find_all('a', href=True) if "wa.me" in a['href'] or "whatsapp.com" in a['href']]
        
        all_links = [a['href'] for a in soup.find_all('a', href=True)]

        
        socials = {"LinkedIn": "", "Twitter": "", "Facebook": "", "Instagram": "", "YouTube": ""}
        for href in all_links:
            if "linkedin.com" in href:
                socials["LinkedIn"] = href
            elif "twitter.com" in href:
                socials["Twitter"] = href
            elif "facebook.com" in href:
                socials["Facebook"] = href
            elif "instagram.com" in href:
                socials["Instagram"] = href
            elif "youtube.com" in href:
                socials["YouTube"] = href

        
        contact_page = next((urljoin(url, link) for link in all_links if "contact" in link.lower()), "N/A")
        careers_page = next((urljoin(url, link) for link in all_links if "career" in link.lower() or "jobs" in link.lower()), "N/A")

        
        partner_links = [link for link in all_links if any(p in link for p in ['.gov', '.org', '.edu', 'partner', 'ngo', 'ministry'])]
        partner_links = list(set(partner_links)) if partner_links else []

       
        address = "N/A"
        for line in text.split("\n"):
            if any(word in line.lower() for word in ['address', 'location', 'hq', 'head office']) and len(line) < 150:
                address = line.strip()
                break

        
        html = res.text.lower()
        tech = []
        for t in ["wordpress", "shopify", "react", "vue", "django", "laravel", "jquery", "bootstrap"]:
            if t in html:
                tech.append(t.capitalize())
        tech_stack = ", ".join(tech) if tech else "Unknown"

        return [{
            "Website": domain,
            "Company Name": company_name,
            "Contact Person": contact_person,
            "Industry": industry,
            "Location": address,
            "Emails": "; ".join(emails) if emails else "N/A",
            "Phones": "; ".join(phones) if phones else "N/A",
            "WhatsApp": "; ".join(whatsapp_links) if whatsapp_links else "N/A",
            "Contact Page": contact_page,
            "Careers Page": careers_page,
            "Tech Stack": tech_stack,
            "LinkedIn": socials["LinkedIn"] or "N/A",
            "Twitter": socials["Twitter"] or "N/A",
            "Facebook": socials["Facebook"] or "N/A",
            "Instagram": socials["Instagram"] or "N/A",
            "YouTube": socials["YouTube"] or "N/A",
            "Gov / Partner Links": "; ".join(partner_links) if partner_links else "N/A"
        }]
    except Exception as e:
        console.print(f"[red] Error scraping {domain}: {e}[/red]")
        return []

def save_to_excel(data, filename="Option 4_Full_ICP.xlsx"):
    
    output_folder = os.path.join(os.getcwd(), "output")
    if not os.path.exists(output_folder):
        os.makedirs(output_folder)

    
    filepath = os.path.join(output_folder, filename)

    if not data:
        console.print("[yellow]⚠️ No data to save.[/yellow]")
        return

    
    console.print(f"[blue]Data to be saved (first 3 entries):[/blue] {data[:3]}")

    
    if os.path.exists(filepath):
       
        wb = load_workbook(filepath)
        ws = wb.active
    else:
        
        wb = Workbook()
        ws = wb.active
        ws.title = "ICP Data"
        
        
        headers = [
            "Website", "Company Name", "Contact Person", "Industry", "Location", "Contact Email", 
            "Phones", "WhatsApp", "Contact Page", "Careers Page", "Tech Stack", "LinkedIn", 
            "Twitter", "Facebook", "Instagram", "YouTube", "Gov / Partner Links"
        ]
        ws.append(headers)

    
    for row in data:
        
        ws.append([
            row.get("Website", "N/A"),
            row.get("Company Name", "N/A"),
            row.get("Contact Person", "N/A"),
            row.get("Industry", "N/A"),
            row.get("Location", "N/A"),
            row.get("Emails", "N/A"),
            row.get("Phones", "N/A"),
            row.get("WhatsApp", "N/A"),
            row.get("Contact Page", "N/A"),
            row.get("Careers Page", "N/A"),
            row.get("Tech Stack", "N/A"),
            row.get("LinkedIn", "N/A"),
            row.get("Twitter", "N/A"),
            row.get("Facebook", "N/A"),
            row.get("Instagram", "N/A"),
            row.get("YouTube", "N/A"),
            row.get("Gov / Partner Links", "N/A")
        ])

    wb.save(filepath)
    console.print(f"\n [green]Saved to {filepath}[/green]")

# ----------------- Option 4: Multi-Domain Search -----------------
def multi_domain_search():
    
    root = tk.Tk()
    root.withdraw()  
    file_path = filedialog.askopenfilename(title="Select Domain List File", filetypes=[("Text files", "*.txt")])
    
    if file_path:
        console.print(f"\n Selected file: {file_path}")
        with open(file_path, 'r') as file:
            domains = file.readlines()
        
        domains = [domain.strip() for domain in domains if domain.strip()] 
        
        all_results = []
        for domain in domains:
            result = extract_icp_from_website(domain)
            all_results.extend(result)
        
       
        save_to_excel(all_results, "Option1_Full_ICP_MultiDomain.xlsx")
    else:
        console.print("[red]No file selected![/red]")


# ----------------- Option 2: SerpAPI Search -----------------
def build_query(industry, location, company_size, tech_stack, website="", search_type="2", linkedin_url=""):
    query = f'"{industry}" "{location}" "{company_size}" "{tech_stack}"'
    if search_type == "1" and website:
        query = f"site:{website} {query}"
    elif search_type == "2":
        query = f'site:linkedin.com/company/ {query}'
    elif search_type == "3" and linkedin_url:
        query = f"site:{linkedin_url} {query}"
    console.print(f"\n🔍 Query: [bold green]{query}[/bold green]\n")
    return query

def google_search(query, start=0, num_results=100):
    url = "https://serpapi.com/search"
    params = {
        "q": query,
        "engine": "google",
        "api_key": SERPAPI_KEY,
        "start": start,
        "num": num_results
    }
    res = requests.get(url, params=params)
    return res.json().get("organic_results", []) if res.status_code == 200 else []

def linkedin_search(query, start=0, num_results=100):
    url = "https://serpapi.com/search"
    params = {
        "q": query,
        "engine": "linkedin",
        "api_key": SERPAPI_KEY,
        "start": start,
        "num": num_results
    }
    res = requests.get(url, params=params)
    return res.json().get("organic_results", []) if res.status_code == 200 else []

def extract_contact_info(url):
    email = ""
    phone = ""
    company_url = ""
    try:
        headers = {"User-Agent": "Mozilla/5.0"}
        html = requests.get(url, headers=headers, timeout=10).text
        soup = BeautifulSoup(html, "html.parser")
        text = soup.get_text()
        email_match = re.findall(r"[a-zA-Z0-9_.+-]+@[a-zA-Z0-9-]+\.[a-zA-Z0-9-.]+", text)
        phone_match = re.findall(r"\+?\d[\d -]{7,}\d", text)
        email = email_match[0] if email_match else ""
        phone = phone_match[0] if phone_match else ""
        company_url = url
    except Exception as e:
        console.print(f" Contact info extraction failed from {url}: {e}")
    return email, phone, company_url

def extract_info_from_result(result):
    title = result.get("title", "")
    link = result.get("link", "")
    snippet = result.get("snippet", "")
    
    company_name = title.split(" |")[0] if "|" in title else title.strip()
    location_keywords = ["USA", "India", "Canada", "UK", "Germany", "Australia", "Singapore"]
    location = "Unknown"
    for keyword in location_keywords:
        if keyword.lower() in title.lower() or keyword.lower() in snippet.lower():
            location = keyword
            break

    industry = "Unknown"
    industry_keywords = ["software", "cybersecurity", "healthcare", "fintech", "education", "consulting", "ai"]
    for keyword in industry_keywords:
        if keyword.lower() in snippet.lower():
            industry = keyword.capitalize()
            break

    tech_stack = "Unknown"
    tech_keywords = ["Python", "Django", "React", "Vue", "Node.js", "Java", "C#", "Ruby", "PHP", "Laravel", "Angular", "AWS", "Azure", "Google Cloud"]
    for tech in tech_keywords:
        if tech.lower() in snippet.lower():
            tech_stack = tech
            break

    company_size = "100+ employees" if "100+" in snippet else "Unknown"

    return company_name, location, industry, company_size, tech_stack, link

def process_result2(result, with_contact_info, without_contact_info):
    company_name, location, industry, company_size, tech_stack, link = extract_info_from_result(result)
    email, phone, company_url = extract_contact_info(link)
    lead = {
        "Company Name": company_name,
        "Contact Person": "",  
        "Industry": industry,  
        "Website/URL": company_url,
        "Location": location,
        "Company Size": company_size,
        "Tech Stack": tech_stack,
        "Contact Email": email,
        "Phone": phone
    }
    if email or phone:
        with_contact_info.append(lead)
    else:
        without_contact_info.append(lead)

def save_to_excel2(data, filename="Option2_General_ICP.xlsx"):
    
    output_folder = os.path.join(os.getcwd(), "output")
    if not os.path.exists(output_folder):
        os.makedirs(output_folder)

    filepath = os.path.join(output_folder, filename)

    if not data:
        console.print("[yellow]⚠️ No data to save.[/yellow]")
        return

    console.print(f"[blue]Data to be saved (first 3 entries):[/blue] {data[:3]}")

    
    if os.path.exists(filepath):
        
        wb = load_workbook(filepath)
        ws = wb.active
    else:
        
        wb = Workbook()
        ws = wb.active
        ws.title = "ICP Data"
        
       
        headers = [
            "Company Name", "Contact Person", "Industry", "Website/URL", "Location", "Company Size", "Tech Stack", 
            "Contact Email", "Phone"
        ]
        ws.append(headers)

    
    for row in data:
        ws.append([
            row.get("Company Name", "N/A"),
            row.get("Contact Person", "N/A"),
            row.get("Industry", "N/A"),
            row.get("Website/URL", "N/A"),
            row.get("Location", "N/A"),
            row.get("Company Size", "N/A"),
            row.get("Tech Stack", "N/A"),
            row.get("Contact Email", "N/A"),
            row.get("Phone", "N/A")
        ])

    wb.save(filepath)
    console.print(f"\n💾 [green]Saved to {filepath}[/green]")

# ----------------- Option 3: ScrapingDog LinkedIn API -----------------
def scrapingdog_linkedin_search(linkedin_url):
    url = "https://api.scrapingdog.com/linkedin"
    company_id = linkedin_url.split("linkedin.com/company/")[-1]
    params = {
        "api_key": SCRAPING_DOG_API_KEY,
        "type": "company",
        "linkId": company_id,
        "private": "false"
    }
    res = requests.get(url, params=params)
    return res.json() if res.status_code == 200 else []

def process_scrapingdog_data(data):
    with_info, without_info = [], []
    if isinstance(data, list):
        for company in data:
            lead = {
                "Company Name": company.get("company_name", ""),
                "Website": company.get("website", ""),
                "Industry": company.get("industry", ""),
                "Size": company.get("company_size", ""),
                "Location": company.get("headquarters", ""),
                "Founded": company.get("founded", ""),
                "Contact Email": company.get("contact_email", ""),
                "Phone": company.get("phone", ""),
                "Employees": "; ".join(emp.get("employee_name", "") for emp in company.get("employees", [])),
                "Employee Positions": "; ".join(emp.get("employee_position", "") for emp in company.get("employees", [])),
                "Profiles": "; ".join(emp.get("employee_profile_url", "") for emp in company.get("employees", [])),
                "Updates": "; ".join(up.get("text", "") for up in company.get("updates", [])),
                "Similar Companies": "; ".join(sim.get("name", "") for sim in company.get("similar_companies", [])),
            }
            if lead["Contact Email"] or lead["Phone"]:
                with_info.append(lead)
            else:
                without_info.append(lead)
    return with_info, without_info

def save_to_excel3(data, filename="TechMantra_Global_ICP.xlsx"):
    
    output_folder = os.path.join(os.getcwd(), "output")
    if not os.path.exists(output_folder):
        os.makedirs(output_folder)

    filepath = os.path.join(output_folder, filename)

    if not data:
        console.print("[yellow]⚠️ No data to save.[/yellow]")
        return

    console.print(f"[blue]Data to be saved (first 3 entries):[/blue] {data[:3]}")

    
    if os.path.exists(filepath):
       
        wb = load_workbook(filepath)
        ws = wb.active
    else:
        
        wb = Workbook()
        ws = wb.active
        ws.title = "ICP Data"
        
        
        headers = [
            "Company Name", "Contact Person", "Industry", "Website/URL", "Location", "Company Size", "Tech Stack", 
            "Contact Email", "Phone", "Employees", "Employee Positions", "Profiles", "Updates", "Similar Companies"
        ]
        ws.append(headers)

    
    for row in data:
        ws.append([
            row.get("Company Name", "N/A"),
            row.get("Contact Person", "N/A"),
            row.get("Industry", "N/A"),
            row.get("Website/URL", "N/A"),
            row.get("Location", "N/A"),
            row.get("Company Size", "N/A"),
            row.get("Tech Stack", "N/A"),
            row.get("Contact Email", "N/A"),
            row.get("Phone", "N/A"),
            row.get("Employees", "N/A"),
            row.get("Employee Positions", "N/A"),
            row.get("Profiles", "N/A"),
            row.get("Updates", "N/A"),
            row.get("Similar Companies", "N/A")
        ])

    wb.save(filepath)
    console.print(f"\n [green]Saved to {filepath}[/green]")

# ----------------- User Input Handler -----------------
def get_user_input():
    console.print("[bold cyan]Choose your search type:[/bold cyan]")
    choice = input("1. Specific Website\n2. General Search (Google & LinkedIn)\n3. Full LinkedIn Company Details (ScrapingDog API)\nChoose an option (1, 2, or 3): ").strip()
    if choice == "1":
        website = input("Enter specific website to scrape (e.g., example.com): ").strip()
        return choice, website
    elif choice == "2":
        industry = input("Industry (e.g., Software Development): ").strip()
        location = input("Location (e.g., Ahmedabad): ").strip()
        company_size = input("Company Size (e.g., 100+ employees): ").strip()
        tech_stack = input("Technology/Stack (e.g., Python, Django): ").strip()
        return choice, industry, location, company_size, tech_stack
    elif choice == "3":
        linkedin_url = input("Enter LinkedIn company URL: ").strip()
        return choice, linkedin_url
    return choice, None

# ----------------- Scraper Launcher -----------------
def scrape_icp_data():
    choice, *inputs = get_user_input()
    all_results = []  

    if choice == "1":
        website = inputs[0]
        result = extract_icp_from_website(website)
        all_results.extend(result)
        
        save_to_excel(all_results, "Option1_Full_ICP.xlsx")
    elif choice == "2":
        industry, location, company_size, tech_stack = inputs
        query = build_query(industry, location, company_size, tech_stack, "", choice)
        with_contact_info = []
        without_contact_info = []
        seen_urls = set()

        with ThreadPoolExecutor(max_workers=5) as executor:
            futures = []
            for start in range(0, 100, 10):  # 2 pages
                futures.append(executor.submit(google_search, query, start))
                futures.append(executor.submit(linkedin_search, query, start))
                
            for future in as_completed(futures):
                results = future.result()
                for r in results:
                    url = r.get("link")
                    if url and url not in seen_urls:
                        seen_urls.add(url)
                        process_result2(r, with_contact_info, without_contact_info)
        
        all_results.extend(with_contact_info)
        all_results.extend(without_contact_info)
        
        save_to_excel2(all_results, "Option2_General_ICP.xlsx")
    elif choice == "3":
        linkedin_url = inputs[0]
        data = scrapingdog_linkedin_search(linkedin_url)
        with_info, without_info = process_scrapingdog_data(data)
        all_results.extend(with_info)
        all_results.extend(without_info)
        
        save_to_excel3(all_results, "Option3_ScrapingDog_ICP.xlsx")

# ----------------- Main Runner -----------------
def main():
    while True:  # 
        
        console.print(
            Panel(
                f"[bold cyan]🚀 Sales Process Automation[/bold cyan]\n"
                f"[bold green]Created by: Dhruv Panchal[/bold green]\n"
                f"[bold magenta]API by: SerpAPI, ScrapingDog[/bold magenta]",
                style="bold cyan", expand=False
            )
        )

        #
        console.print("\n[bold cyan]Choose your search type:[/bold cyan]")
        console.print("[bold green]1.[/bold green] Specific Website (No API)")
        console.print("[bold green]2.[/bold green] General Search with ICP (Slow More Result++) (Google & LinkedIn)")
        console.print("[bold green]3.[/bold green] Full LinkedIn Company Details (ScrapingDog API)")
        console.print("[bold green]4.[/bold green] Multi-Domain Search (No API)")
        console.print("[bold red]5.[/bold red] Exit")

        
        choice = Prompt.ask(
            Text("Please select an option", style="bold green"),
            choices=["1", "2", "3", "4", "5"],
            default="1",
            show_choices=True
        )

        if choice == "1":
            console.print("\n[bold blue]You chose: Specific Website (No API)[/bold blue]")
            website = Prompt.ask(Text("Enter the specific website to scrape (e.g., example.com):", style="bold yellow"))
            result = extract_icp_from_website(website)
            save_to_excel(result, "Option1_Full_ICP.xlsx")
        
        elif choice == "2":
            console.print("\n[bold blue]You chose: General Search with ICP (Google & LinkedIn)[/bold blue]")
            industry = Prompt.ask(Text("Industry (e.g., Software Development):", style="bold yellow"))
            location = Prompt.ask(Text("Location (e.g., Ahmedabad):", style="bold yellow"))
            company_size = Prompt.ask(Text("Company Size (e.g., 100+ employees):", style="bold yellow"))
            tech_stack = Prompt.ask(Text("Technology/Stack (e.g., Python, Django):", style="bold yellow"))
            
            query = build_query(industry, location, company_size, tech_stack, "", choice)
            with_contact_info = []
            without_contact_info = []
            seen_urls = set()

            with ThreadPoolExecutor(max_workers=5) as executor:
                futures = []
                for start in range(0, 100, 10):  #2 pages
                    futures.append(executor.submit(google_search, query, start))
                    futures.append(executor.submit(linkedin_search, query, start))

                for future in as_completed(futures):
                    results = future.result()
                    for r in results:
                        url = r.get("link")
                        if url and url not in seen_urls:
                            seen_urls.add(url)
                            process_result2(r, with_contact_info, without_contact_info)

            result = with_contact_info + without_contact_info
            save_to_excel2(result, "Option2_General_ICP.xlsx")
        
        elif choice == "3":
            console.print("\n[bold blue]You chose: Full LinkedIn Company Details (ScrapingDog API)[/bold blue]")
            linkedin_url = Prompt.ask(Text("Enter LinkedIn company URL:", style="bold yellow"))
            data = scrapingdog_linkedin_search(linkedin_url)
            with_info, without_info = process_scrapingdog_data(data)
            result = with_info + without_info
            save_to_excel3(result, "Option3_ScrapingDog_ICP.xlsx")
        
        elif choice == "4":
            console.print("\n[bold blue]You chose: Multi-Domain Search (No API)[/bold blue]")
            multi_domain_search()  
        
        elif choice == "5":
            console.print("\n[bold red]Exiting the program...[/bold red]")
            break  
        
        else:
            console.print("\n[bold red]Invalid option! Please choose a valid option (1, 2, 3, 4, or 5).[/bold red]")
            continue

if __name__ == "__main__":
    main()




